from flask import Blueprint, render_template, request, send_file, session
from models.report import Report
from utils.decorators import login_required, permission_required

from io import BytesIO
from datetime import datetime
from urllib.parse import urlencode

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer
)


report_bp = Blueprint('reports', __name__, url_prefix='/reports')


def _clean_filter(value):
    """
    Limpiar filtros vacíos.
    """
    if value is None:
        return None

    value = str(value).strip()

    if value == '':
        return None

    return value


def _get_sales_filters_from_request():
    """
    Obtener filtros del reporte de ventas.
    """
    return {
        'fecha_inicio': _clean_filter(request.args.get('fecha_inicio')),
        'fecha_fin': _clean_filter(request.args.get('fecha_fin')),
        'codigo_usuario': _clean_filter(request.args.get('codigo_usuario')),
        'id_tipo_pago': _clean_filter(request.args.get('id_tipo_pago')),
        'estado': _clean_filter(request.args.get('estado')) or 'todos',
        'search': _clean_filter(request.args.get('search'))
    }


def _get_profit_filters_from_request():
    """
    Obtener filtros del reporte de utilidad.
    Por defecto se muestran solo facturas activas porque representan utilidad real.
    """
    return {
        'fecha_inicio': _clean_filter(request.args.get('fecha_inicio')),
        'fecha_fin': _clean_filter(request.args.get('fecha_fin')),
        'codigo_usuario': _clean_filter(request.args.get('codigo_usuario')),
        'codigo_producto': _clean_filter(request.args.get('codigo_producto')),
        'id_tipo_pago': _clean_filter(request.args.get('id_tipo_pago')),
        'estado': _clean_filter(request.args.get('estado')) or 'activa',
        'search': _clean_filter(request.args.get('search'))
    }


def _get_selected_fields():
    """
    Obtener campos seleccionados del reporte de ventas.
    Si no se selecciona ninguno, se usan campos por defecto.
    """
    available_fields = Report.get_available_fields()
    selected_fields = request.args.getlist('fields')

    selected_fields = [
        field for field in selected_fields
        if field in available_fields
    ]

    if not selected_fields:
        selected_fields = Report.get_default_fields()

    return selected_fields


def _format_value(value, field_name=None):
    """
    Formatear valores para mostrar en HTML, PDF y Excel.
    """
    if value is None:
        return '-'

    if field_name == 'fecha_hora':
        try:
            return value.strftime('%d/%m/%Y %H:%M:%S')
        except Exception:
            return str(value)

    money_fields = [
        'monto_total',
        'monto_final',
        'total_bruto',
        'total_final_cobrado',
        'costo_total_productos',
        'utilidad',
        'descuento_total',
        'monto_promocion',
        'monto_pagado'
    ]

    if field_name in money_fields:
        try:
            return f'Bs {float(value):.2f}'
        except Exception:
            return str(value)

    if field_name == 'margen_utilidad':
        try:
            return f'{float(value):.2f}%'
        except Exception:
            return str(value)

    if field_name == 'pagado':
        return 'Sí' if value else 'No'

    if field_name == 'estado':
        return 'Activa' if value else 'Anulada'

    return str(value)


def _build_query_string(filters, selected_fields=None):
    """
    Construir querystring para botones de exportación.
    """
    query_data = {}

    for key, value in filters.items():
        if value:
            query_data[key] = value

    if selected_fields:
        query_data['fields'] = selected_fields

    return urlencode(query_data, doseq=True)


# =========================================================
# PDF - REPORTE DE VENTAS EXISTENTE
# =========================================================

def _build_sales_pdf(rows, summary, selected_fields, filters):
    """
    Generar PDF del reporte de ventas usando ReportLab.
    """
    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=1.2 * cm,
        leftMargin=1.2 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        name='TitleCenter',
        parent=styles['Title'],
        alignment=TA_CENTER,
        fontSize=16,
        leading=20,
        spaceAfter=10
    )

    subtitle_style = ParagraphStyle(
        name='SubtitleCenter',
        parent=styles['Normal'],
        alignment=TA_CENTER,
        fontSize=9,
        textColor=colors.HexColor('#555555'),
        spaceAfter=12
    )

    normal_style = ParagraphStyle(
        name='NormalReport',
        parent=styles['Normal'],
        fontSize=8,
        leading=10,
        alignment=TA_LEFT
    )

    small_style = ParagraphStyle(
        name='SmallReport',
        parent=styles['Normal'],
        fontSize=7,
        leading=9
    )

    elements = []

    elements.append(Paragraph('MiniMarket QuickStore', title_style))
    elements.append(Paragraph('Reporte de Ventas', title_style))

    generated_at = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    elements.append(Paragraph(f'Generado el: {generated_at}', subtitle_style))

    filtros_texto = []

    if filters.get('fecha_inicio'):
        filtros_texto.append(f"Fecha inicio: {filters['fecha_inicio']}")

    if filters.get('fecha_fin'):
        filtros_texto.append(f"Fecha fin: {filters['fecha_fin']}")

    if filters.get('codigo_usuario'):
        filtros_texto.append(f"Usuario: {filters['codigo_usuario']}")

    if filters.get('id_tipo_pago'):
        filtros_texto.append(f"Tipo de pago ID: {filters['id_tipo_pago']}")

    if filters.get('estado') and filters.get('estado') != 'todos':
        filtros_texto.append(f"Estado: {filters['estado']}")

    if filters.get('search'):
        filtros_texto.append(f"Búsqueda: {filters['search']}")

    if not filtros_texto:
        filtros_texto.append('Sin filtros aplicados')

    elements.append(Paragraph('Filtros: ' + ' | '.join(filtros_texto), normal_style))
    elements.append(Spacer(1, 0.25 * cm))

    resumen_data = [
        ['Total registros', 'Ventas activas', 'Ventas anuladas', 'Total activo', 'Total general'],
        [
            str(summary['total_registros']),
            str(summary['ventas_activas']),
            str(summary['ventas_anuladas']),
            f"Bs {float(summary['total_activo']):.2f}",
            f"Bs {float(summary['total_general']):.2f}"
        ]
    ]

    resumen_table = Table(resumen_data, repeatRows=1)
    resumen_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f2937')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#cccccc')),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f3f4f6')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
    ]))

    elements.append(resumen_table)
    elements.append(Spacer(1, 0.4 * cm))

    available_fields = Report.get_available_fields()

    header = [
        Paragraph(f'<b>{available_fields[field]}</b>', small_style)
        for field in selected_fields
    ]

    data = [header]

    for row in rows:
        row_data = []

        for field in selected_fields:
            formatted_value = _format_value(row.get(field), field)
            row_data.append(Paragraph(formatted_value, small_style))

        data.append(row_data)

    if len(data) == 1:
        data.append([
            Paragraph('No existen datos para los filtros seleccionados.', small_style)
        ] + ['' for _ in selected_fields[1:]])

    table = Table(data, repeatRows=1)

    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#111827')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#d1d5db')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [
            colors.white,
            colors.HexColor('#f9fafb')
        ]),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
    ]))

    elements.append(table)

    doc.build(elements)

    buffer.seek(0)

    return buffer


# =========================================================
# PDF - CU17 REPORTE DE UTILIDAD
# =========================================================

def _build_profit_pdf(rows, summary, filters):
    """
    Generar PDF del reporte de utilidad.
    """
    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=1.0 * cm,
        leftMargin=1.0 * cm,
        topMargin=1.0 * cm,
        bottomMargin=1.0 * cm
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        name='TitleCenterProfit',
        parent=styles['Title'],
        alignment=TA_CENTER,
        fontSize=15,
        leading=18,
        spaceAfter=8
    )

    subtitle_style = ParagraphStyle(
        name='SubtitleCenterProfit',
        parent=styles['Normal'],
        alignment=TA_CENTER,
        fontSize=8,
        textColor=colors.HexColor('#555555'),
        spaceAfter=10
    )

    normal_style = ParagraphStyle(
        name='NormalProfit',
        parent=styles['Normal'],
        fontSize=7,
        leading=9,
        alignment=TA_LEFT
    )

    small_style = ParagraphStyle(
        name='SmallProfit',
        parent=styles['Normal'],
        fontSize=6.5,
        leading=8
    )

    elements = []

    elements.append(Paragraph('MiniMarket QuickStore', title_style))
    elements.append(Paragraph('CU17 - Reporte de Utilidad', title_style))

    generated_at = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    elements.append(Paragraph(f'Generado el: {generated_at}', subtitle_style))

    filtros_texto = []

    if filters.get('fecha_inicio'):
        filtros_texto.append(f"Fecha inicio: {filters['fecha_inicio']}")

    if filters.get('fecha_fin'):
        filtros_texto.append(f"Fecha fin: {filters['fecha_fin']}")

    if filters.get('codigo_usuario'):
        filtros_texto.append(f"Cajero: {filters['codigo_usuario']}")

    if filters.get('codigo_producto'):
        filtros_texto.append(f"Producto: {filters['codigo_producto']}")

    if filters.get('id_tipo_pago'):
        filtros_texto.append(f"Tipo de pago ID: {filters['id_tipo_pago']}")

    if filters.get('estado') and filters.get('estado') != 'todos':
        filtros_texto.append(f"Estado: {filters['estado']}")

    if filters.get('search'):
        filtros_texto.append(f"Búsqueda: {filters['search']}")

    if not filtros_texto:
        filtros_texto.append('Sin filtros aplicados')

    elements.append(Paragraph('Filtros: ' + ' | '.join(filtros_texto), normal_style))
    elements.append(Spacer(1, 0.2 * cm))

    resumen_data = [
        [
            'Facturas',
            'Total bruto',
            'Final cobrado',
            'Costo vendido',
            'Utilidad',
            'Margen',
            'Desc./Promo.'
        ],
        [
            str(summary['total_facturas']),
            f"Bs {float(summary['total_bruto']):.2f}",
            f"Bs {float(summary['total_final_cobrado']):.2f}",
            f"Bs {float(summary['costo_total_productos']):.2f}",
            f"Bs {float(summary['utilidad_total']):.2f}",
            f"{float(summary['margen_global']):.2f}%",
            f"Bs {float(summary['descuento_total']):.2f}"
        ]
    ]

    resumen_table = Table(resumen_data, repeatRows=1)
    resumen_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#14532d')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#cccccc')),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f0fdf4')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
    ]))

    elements.append(resumen_table)
    elements.append(Spacer(1, 0.3 * cm))

    header = [
        'Factura',
        'Fecha',
        'Cliente',
        'Cajero',
        'Bruto',
        'Final',
        'Costo',
        'Utilidad',
        'Margen',
        'Pago',
        'Estado'
    ]

    data = [[Paragraph(f'<b>{col}</b>', small_style) for col in header]]

    for row in rows:
        data.append([
            Paragraph(str(row.get('codigo_factura') or '-'), small_style),
            Paragraph(_format_value(row.get('fecha_hora'), 'fecha_hora'), small_style),
            Paragraph(str(row.get('cliente') or '-'), small_style),
            Paragraph(str(row.get('usuario') or '-'), small_style),
            Paragraph(_format_value(row.get('total_bruto'), 'total_bruto'), small_style),
            Paragraph(_format_value(row.get('total_final_cobrado'), 'total_final_cobrado'), small_style),
            Paragraph(_format_value(row.get('costo_total_productos'), 'costo_total_productos'), small_style),
            Paragraph(_format_value(row.get('utilidad'), 'utilidad'), small_style),
            Paragraph(_format_value(row.get('margen_utilidad'), 'margen_utilidad'), small_style),
            Paragraph(str(row.get('tipo_pago') or '-'), small_style),
            Paragraph(_format_value(row.get('estado'), 'estado'), small_style)
        ])

    if len(data) == 1:
        data.append([
            Paragraph('No existen datos para los filtros seleccionados.', small_style)
        ] + ['' for _ in header[1:]])

    table = Table(data, repeatRows=1)

    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#111827')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 6.5),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#d1d5db')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [
            colors.white,
            colors.HexColor('#f9fafb')
        ]),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
    ]))

    elements.append(table)

    doc.build(elements)

    buffer.seek(0)

    return buffer


# =========================================================
# EXCEL - CU17 REPORTE DE UTILIDAD
# =========================================================

def _build_profit_excel(rows, summary, filters):
    """
    Generar Excel del reporte de utilidad.
    Requiere:
    pip install openpyxl
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Reporte Utilidad'

    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='14532D')
    table_header_fill = PatternFill('solid', fgColor='111827')
    border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    sheet['A1'] = 'MiniMarket QuickStore'
    sheet['A1'].font = title_font

    sheet['A2'] = 'CU17 - Reporte de Utilidad'
    sheet['A2'].font = title_font

    sheet['A3'] = f'Generado el: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}'

    sheet['A5'] = 'Filtros aplicados'
    sheet['A5'].font = Font(bold=True)

    filter_labels = [
        ('Fecha inicio', filters.get('fecha_inicio') or '-'),
        ('Fecha fin', filters.get('fecha_fin') or '-'),
        ('Cajero', filters.get('codigo_usuario') or '-'),
        ('Producto', filters.get('codigo_producto') or '-'),
        ('Tipo de pago', filters.get('id_tipo_pago') or '-'),
        ('Estado', filters.get('estado') or '-'),
        ('Búsqueda', filters.get('search') or '-')
    ]

    start_filter_row = 6

    for index, item in enumerate(filter_labels, start=start_filter_row):
        sheet[f'A{index}'] = item[0]
        sheet[f'B{index}'] = item[1]
        sheet[f'A{index}'].font = Font(bold=True)

    summary_row = 15

    summary_headers = [
        'Facturas',
        'Facturas activas',
        'Facturas anuladas',
        'Total bruto',
        'Final cobrado',
        'Costo vendido',
        'Utilidad',
        'Margen %',
        'Desc./Promo.'
    ]

    summary_values = [
        summary['total_facturas'],
        summary['facturas_activas'],
        summary['facturas_anuladas'],
        float(summary['total_bruto']),
        float(summary['total_final_cobrado']),
        float(summary['costo_total_productos']),
        float(summary['utilidad_total']),
        float(summary['margen_global']),
        float(summary['descuento_total'])
    ]

    for col_index, header in enumerate(summary_headers, start=1):
        cell = sheet.cell(row=summary_row, column=col_index, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    for col_index, value in enumerate(summary_values, start=1):
        cell = sheet.cell(row=summary_row + 1, column=col_index, value=value)
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    table_row = 19

    headers = [
        'Factura',
        'Fecha',
        'Cliente',
        'CI Cliente',
        'Cajero',
        'Tipo de Pago',
        'Total Bruto',
        'Total Final Cobrado',
        'Costo Total Productos',
        'Utilidad',
        'Margen %',
        'Descuento/Promoción',
        'Promoción',
        'Descuento',
        'Estado',
        'Referencia'
    ]

    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=table_row, column=col_index, value=header)
        cell.font = header_font
        cell.fill = table_header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    current_row = table_row + 1

    for row in rows:
        values = [
            row.get('codigo_factura'),
            _format_value(row.get('fecha_hora'), 'fecha_hora'),
            row.get('cliente'),
            row.get('cliente_ci'),
            row.get('usuario'),
            row.get('tipo_pago'),
            float(row.get('total_bruto') or 0),
            float(row.get('total_final_cobrado') or 0),
            float(row.get('costo_total_productos') or 0),
            float(row.get('utilidad') or 0),
            float(row.get('margen_utilidad') or 0),
            float(row.get('descuento_total') or 0),
            row.get('promociones') or '-',
            row.get('descuento') or '-',
            _format_value(row.get('estado'), 'estado'),
            row.get('referencia_externa') or '-'
        ]

        for col_index, value in enumerate(values, start=1):
            cell = sheet.cell(row=current_row, column=col_index, value=value)
            cell.border = border

        current_row += 1

    if not rows:
        sheet.cell(
            row=current_row,
            column=1,
            value='No existen datos para los filtros seleccionados.'
        )

    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        sheet.column_dimensions[column_letter].width = 18

    sheet.column_dimensions['C'].width = 28
    sheet.column_dimensions['E'].width = 24
    sheet.column_dimensions['M'].width = 30
    sheet.column_dimensions['N'].width = 24
    sheet.column_dimensions['P'].width = 28

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    return buffer


# =========================================================
# RUTAS - REPORTE DE VENTAS EXISTENTE
# =========================================================

@report_bp.route('/sales')
@login_required
@permission_required('report_sales_read')
def sales_report():
    """
    Pantalla del reporte de ventas.
    """
    filters = _get_sales_filters_from_request()
    selected_fields = _get_selected_fields()

    rows = Report.get_sales_report(
        fecha_inicio=filters['fecha_inicio'],
        fecha_fin=filters['fecha_fin'],
        codigo_usuario=filters['codigo_usuario'],
        id_tipo_pago=filters['id_tipo_pago'],
        estado=filters['estado'],
        search=filters['search']
    )

    summary = Report.calculate_summary(rows)

    users = Report.get_users_for_filter()
    payment_types = Report.get_payment_types_for_filter()

    pdf_query = _build_query_string(filters, selected_fields)

    return render_template(
        'reports/sales.html',
        rows=rows,
        summary=summary,
        users=users,
        payment_types=payment_types,
        filters=filters,
        available_fields=Report.get_available_fields(),
        selected_fields=selected_fields,
        pdf_query=pdf_query,
        format_value=_format_value
    )


@report_bp.route('/sales/pdf')
@login_required
@permission_required('report_sales_pdf')
def sales_report_pdf():
    """
    Descargar el reporte de ventas en PDF.
    """
    filters = _get_sales_filters_from_request()
    selected_fields = _get_selected_fields()

    rows = Report.get_sales_report(
        fecha_inicio=filters['fecha_inicio'],
        fecha_fin=filters['fecha_fin'],
        codigo_usuario=filters['codigo_usuario'],
        id_tipo_pago=filters['id_tipo_pago'],
        estado=filters['estado'],
        search=filters['search']
    )

    summary = Report.calculate_summary(rows)

    pdf_buffer = _build_sales_pdf(
        rows=rows,
        summary=summary,
        selected_fields=selected_fields,
        filters=filters
    )

    try:
        Report.register_report_action(
            codigo_usuario=session.get('user_id'),
            accion='DESCARGA DE REPORTE DE VENTAS PDF',
            detalle=f'Se descargó reporte de ventas con {summary["total_registros"]} registros.'
        )
    except Exception:
        pass

    filename = f'reporte_ventas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'

    return send_file(
        pdf_buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/pdf'
    )


# =========================================================
# RUTAS - CU17 REPORTE DE UTILIDAD
# =========================================================

@report_bp.route('/profit')
@login_required
@permission_required('report_profit_read')
def profit_report():
    """
    Pantalla del reporte de utilidad.
    """
    filters = _get_profit_filters_from_request()

    rows = Report.get_profit_report(
        fecha_inicio=filters['fecha_inicio'],
        fecha_fin=filters['fecha_fin'],
        codigo_usuario=filters['codigo_usuario'],
        codigo_producto=filters['codigo_producto'],
        id_tipo_pago=filters['id_tipo_pago'],
        estado=filters['estado'],
        search=filters['search']
    )

    summary = Report.calculate_profit_summary(rows)

    users = Report.get_users_for_filter()
    payment_types = Report.get_payment_types_for_filter()
    products = Report.get_products_for_filter()

    export_query = _build_query_string(filters)

    return render_template(
        'reports/profit.html',
        rows=rows,
        summary=summary,
        users=users,
        payment_types=payment_types,
        products=products,
        filters=filters,
        export_query=export_query,
        format_value=_format_value
    )


@report_bp.route('/profit/pdf')
@login_required
@permission_required('report_profit_pdf')
def profit_report_pdf():
    """
    Descargar reporte de utilidad en PDF.
    """
    filters = _get_profit_filters_from_request()

    rows = Report.get_profit_report(
        fecha_inicio=filters['fecha_inicio'],
        fecha_fin=filters['fecha_fin'],
        codigo_usuario=filters['codigo_usuario'],
        codigo_producto=filters['codigo_producto'],
        id_tipo_pago=filters['id_tipo_pago'],
        estado=filters['estado'],
        search=filters['search']
    )

    summary = Report.calculate_profit_summary(rows)

    pdf_buffer = _build_profit_pdf(
        rows=rows,
        summary=summary,
        filters=filters
    )

    try:
        Report.register_report_action(
            codigo_usuario=session.get('user_id'),
            accion='DESCARGA DE REPORTE DE UTILIDAD PDF',
            detalle=f'Se descargó reporte de utilidad con {summary["total_facturas"]} facturas.'
        )
    except Exception:
        pass

    filename = f'reporte_utilidad_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'

    return send_file(
        pdf_buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/pdf'
    )


@report_bp.route('/profit/excel')
@login_required
@permission_required('report_profit_excel')
def profit_report_excel():
    """
    Descargar reporte de utilidad en Excel.
    """
    filters = _get_profit_filters_from_request()

    rows = Report.get_profit_report(
        fecha_inicio=filters['fecha_inicio'],
        fecha_fin=filters['fecha_fin'],
        codigo_usuario=filters['codigo_usuario'],
        codigo_producto=filters['codigo_producto'],
        id_tipo_pago=filters['id_tipo_pago'],
        estado=filters['estado'],
        search=filters['search']
    )

    summary = Report.calculate_profit_summary(rows)

    excel_buffer = _build_profit_excel(
        rows=rows,
        summary=summary,
        filters=filters
    )

    try:
        Report.register_report_action(
            codigo_usuario=session.get('user_id'),
            accion='DESCARGA DE REPORTE DE UTILIDAD EXCEL',
            detalle=f'Se descargó reporte de utilidad Excel con {summary["total_facturas"]} facturas.'
        )
    except Exception:
        pass

    filename = f'reporte_utilidad_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    return send_file(
        excel_buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )