from io import BytesIO
from datetime import datetime
from html import escape

from flask import Blueprint, render_template, redirect, url_for, flash, session, send_file, Response

from models.managerial_dashboard import ManagerialDashboard
from utils.decorators import login_required, admin_required
from utils.gemini_client import GeminiClient
from utils.email_service import EmailService


managerial_dashboard_bp = Blueprint(
    'managerial_dashboard',
    __name__,
    url_prefix='/reports/managerial-dashboard'
)


def _get_report_filename(extension):
    """
    Genera nombre de archivo para exportaciones del Dashboard Gerencial.
    """
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return f'dashboard_gerencial_quickstore_{timestamp}.{extension}'


def _get_indicator_rows(indicators):
    """
    Convierte los indicadores del dashboard en filas listas para PDF, Excel y HTML.
    """
    formatted = indicators.get('formatted', {})
    payment_method = indicators.get('metodo_pago_mas_usado', {})

    return [
        ('Ventas del día', formatted.get('ventas_dia', 'Bs 0.00')),
        ('Ventas del mes', formatted.get('ventas_mes', 'Bs 0.00')),
        ('Utilidad del mes', formatted.get('utilidad_mes', 'Bs 0.00')),
        ('Productos con stock bajo', indicators.get('productos_stock_bajo', 0)),
        ('Productos agotados', indicators.get('productos_agotados', 0)),
        ('Productos próximos a vencer', indicators.get('productos_proximos_vencer', 0)),
        ('Compras del mes', formatted.get('compras_mes', 'Bs 0.00')),
        ('Pérdidas por notas de salida', formatted.get('perdidas_notas_salida', 'Bs 0.00')),
        ('Cajas abiertas', indicators.get('cajas_abiertas', 0)),
        (
            'Método de pago más usado',
            f"{payment_method.get('metodo_pago', 'Sin registros')} "
            f"({payment_method.get('cantidad', 0)} uso/s, "
            f"{formatted.get('metodo_pago_total', 'Bs 0.00')})"
        )
    ]


def _get_sales_rows(sales_by_day):
    """
    Convierte las ventas por día en filas para exportación.
    """
    rows = []

    for item in sales_by_day:
        total = item.get('total', 0)

        try:
            total_formatted = f"Bs {float(total):.2f}"
        except Exception:
            total_formatted = 'Bs 0.00'

        rows.append((
            item.get('day', '-'),
            item.get('date', '-'),
            total_formatted
        ))

    return rows


def _register_dashboard_action(accion, detalle):
    """
    Registra acciones del dashboard gerencial en bitácora.
    No rompe la operación si la bitácora falla.
    """
    try:
        ManagerialDashboard.register_action(
            codigo_usuario=session.get('user_id'),
            accion=accion,
            detalle=detalle
        )
    except Exception:
        pass


@managerial_dashboard_bp.route('/')
@login_required
@admin_required
def managerial_dashboard_page():
    """
    Pantalla principal del CU19 - Dashboard Gerencial Inteligente.
    Solo administradores.
    """
    dashboard_data = ManagerialDashboard.get_dashboard_data()

    _register_dashboard_action(
        accion='CONSULTA DASHBOARD GERENCIAL',
        detalle='El administrador consultó el Dashboard Gerencial Inteligente.'
    )

    return render_template(
        'reports/managerial_dashboard.html',
        indicators=dashboard_data['indicators'],
        sales_by_day=dashboard_data['sales_by_day'],
        chart_labels=dashboard_data['chart_labels'],
        chart_values=dashboard_data['chart_values'],
        has_month_sales=dashboard_data['has_month_sales']
    )


@managerial_dashboard_bp.route('/send-ai-email', methods=['POST'])
@login_required
@admin_required
def send_ai_email():
    """
    Generar análisis gerencial con Gemini IA y enviarlo por correo.
    Solo administradores.
    """
    try:
        dashboard_data = ManagerialDashboard.get_dashboard_data()
        indicators = dashboard_data['indicators']

        analysis = GeminiClient.generate_managerial_analysis(indicators)
        sent_to = EmailService.send_managerial_analysis(analysis, indicators)

        _register_dashboard_action(
            accion='ENVÍO ANÁLISIS GERENCIAL IA POR CORREO',
            detalle=f'Se envió análisis gerencial IA al correo {sent_to}.'
        )

        flash(f'Análisis gerencial IA enviado correctamente a {sent_to}.', 'success')

    except Exception as e:
        _register_dashboard_action(
            accion='ERROR ENVÍO ANÁLISIS GERENCIAL IA POR CORREO',
            detalle=f'Error al enviar análisis gerencial IA: {str(e)}'
        )

        flash(f'No se pudo enviar el análisis gerencial: {str(e)}', 'danger')

    return redirect(url_for('managerial_dashboard.managerial_dashboard_page'))


@managerial_dashboard_bp.route('/pdf')
@login_required
@admin_required
def export_pdf():
    """
    Descargar reporte gerencial en PDF.
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        dashboard_data = ManagerialDashboard.get_dashboard_data()
        indicators = dashboard_data['indicators']
        sales_by_day = dashboard_data['sales_by_day']

        indicator_rows = _get_indicator_rows(indicators)
        sales_rows = _get_sales_rows(sales_by_day)

        buffer = BytesIO()

        document = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=32,
            leftMargin=32,
            topMargin=32,
            bottomMargin=32
        )

        styles = getSampleStyleSheet()
        elements = []

        title = Paragraph('Dashboard Gerencial QuickStore', styles['Title'])
        subtitle = Paragraph(
            f"Reporte generado el {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
            styles['Normal']
        )

        elements.append(title)
        elements.append(Spacer(1, 10))
        elements.append(subtitle)
        elements.append(Spacer(1, 18))

        elements.append(Paragraph('Indicadores principales', styles['Heading2']))
        elements.append(Spacer(1, 8))

        indicator_table_data = [['Indicador', 'Valor']]

        for label, value in indicator_rows:
            indicator_table_data.append([str(label), str(value)])

        indicator_table = Table(indicator_table_data, colWidths=[230, 260])
        indicator_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f5db8')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),
        ]))

        elements.append(indicator_table)
        elements.append(Spacer(1, 20))

        elements.append(Paragraph('Ventas por día del mes actual', styles['Heading2']))
        elements.append(Spacer(1, 8))

        sales_table_data = [['Día', 'Fecha', 'Total vendido']]

        for day, date, total in sales_rows:
            sales_table_data.append([str(day), str(date), str(total)])

        sales_table = Table(sales_table_data, colWidths=[80, 190, 220])
        sales_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f5db8')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),
        ]))

        elements.append(sales_table)

        document.build(elements)
        buffer.seek(0)

        _register_dashboard_action(
            accion='DESCARGA DASHBOARD GERENCIAL PDF',
            detalle='El administrador descargó el Dashboard Gerencial en PDF.'
        )

        return send_file(
            buffer,
            as_attachment=True,
            download_name=_get_report_filename('pdf'),
            mimetype='application/pdf'
        )

    except Exception as e:
        _register_dashboard_action(
            accion='ERROR DESCARGA DASHBOARD GERENCIAL PDF',
            detalle=f'Error al descargar PDF del Dashboard Gerencial: {str(e)}'
        )

        flash(f'No se pudo generar el PDF: {str(e)}', 'danger')
        return redirect(url_for('managerial_dashboard.managerial_dashboard_page'))


@managerial_dashboard_bp.route('/excel')
@login_required
@admin_required
def export_excel():
    """
    Descargar reporte gerencial en Excel.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        dashboard_data = ManagerialDashboard.get_dashboard_data()
        indicators = dashboard_data['indicators']
        sales_by_day = dashboard_data['sales_by_day']

        indicator_rows = _get_indicator_rows(indicators)
        sales_rows = _get_sales_rows(sales_by_day)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Dashboard Gerencial'

        header_fill = PatternFill('solid', fgColor='0F5DB8')
        header_font = Font(color='FFFFFF', bold=True)
        title_font = Font(size=16, bold=True, color='0D3F73')
        subtitle_font = Font(size=10, color='64748B')
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        sheet['A1'] = 'Dashboard Gerencial QuickStore'
        sheet['A1'].font = title_font
        sheet.merge_cells('A1:C1')

        sheet['A2'] = f"Reporte generado el {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        sheet['A2'].font = subtitle_font
        sheet.merge_cells('A2:C2')

        current_row = 4

        sheet[f'A{current_row}'] = 'Indicadores principales'
        sheet[f'A{current_row}'].font = Font(size=13, bold=True, color='0D3F73')
        current_row += 1

        sheet[f'A{current_row}'] = 'Indicador'
        sheet[f'B{current_row}'] = 'Valor'

        for cell in [sheet[f'A{current_row}'], sheet[f'B{current_row}']]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        current_row += 1

        for label, value in indicator_rows:
            sheet[f'A{current_row}'] = str(label)
            sheet[f'B{current_row}'] = str(value)

            sheet[f'A{current_row}'].border = thin_border
            sheet[f'B{current_row}'].border = thin_border

            sheet[f'A{current_row}'].alignment = Alignment(vertical='center')
            sheet[f'B{current_row}'].alignment = Alignment(vertical='center')

            current_row += 1

        current_row += 2

        sheet[f'A{current_row}'] = 'Ventas por día del mes actual'
        sheet[f'A{current_row}'].font = Font(size=13, bold=True, color='0D3F73')
        current_row += 1

        sheet[f'A{current_row}'] = 'Día'
        sheet[f'B{current_row}'] = 'Fecha'
        sheet[f'C{current_row}'] = 'Total vendido'

        for cell in [sheet[f'A{current_row}'], sheet[f'B{current_row}'], sheet[f'C{current_row}']]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        current_row += 1

        for day, date, total in sales_rows:
            sheet[f'A{current_row}'] = str(day)
            sheet[f'B{current_row}'] = str(date)
            sheet[f'C{current_row}'] = str(total)

            for column in ['A', 'B', 'C']:
                sheet[f'{column}{current_row}'].border = thin_border
                sheet[f'{column}{current_row}'].alignment = Alignment(vertical='center')

            current_row += 1

        for column_index in range(1, 4):
            column_letter = get_column_letter(column_index)
            sheet.column_dimensions[column_letter].width = 28

        sheet.column_dimensions['A'].width = 34
        sheet.column_dimensions['B'].width = 28
        sheet.column_dimensions['C'].width = 24

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        _register_dashboard_action(
            accion='DESCARGA DASHBOARD GERENCIAL EXCEL',
            detalle='El administrador descargó el Dashboard Gerencial en Excel.'
        )

        return send_file(
            buffer,
            as_attachment=True,
            download_name=_get_report_filename('xlsx'),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        _register_dashboard_action(
            accion='ERROR DESCARGA DASHBOARD GERENCIAL EXCEL',
            detalle=f'Error al descargar Excel del Dashboard Gerencial: {str(e)}'
        )

        flash(f'No se pudo generar el Excel: {str(e)}', 'danger')
        return redirect(url_for('managerial_dashboard.managerial_dashboard_page'))


@managerial_dashboard_bp.route('/html')
@login_required
@admin_required
def export_html():
    """
    Descargar reporte gerencial en HTML.
    """
    try:
        dashboard_data = ManagerialDashboard.get_dashboard_data()
        indicators = dashboard_data['indicators']
        sales_by_day = dashboard_data['sales_by_day']

        indicator_rows = _get_indicator_rows(indicators)
        sales_rows = _get_sales_rows(sales_by_day)

        indicator_html_rows = ''

        for label, value in indicator_rows:
            indicator_html_rows += f"""
                <tr>
                    <td>{escape(str(label))}</td>
                    <td>{escape(str(value))}</td>
                </tr>
            """

        sales_html_rows = ''

        for day, date, total in sales_rows:
            sales_html_rows += f"""
                <tr>
                    <td>{escape(str(day))}</td>
                    <td>{escape(str(date))}</td>
                    <td>{escape(str(total))}</td>
                </tr>
            """

        generated_at = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

        html_content = f"""<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <title>Dashboard Gerencial QuickStore</title>
    <style>
        body {{
            font-family: Arial, sans-serif;
            background: #f4f7fb;
            color: #0f172a;
            margin: 0;
            padding: 28px;
        }}

        .container {{
            max-width: 1000px;
            margin: 0 auto;
            background: #ffffff;
            border-radius: 18px;
            overflow: hidden;
            box-shadow: 0 12px 30px rgba(15, 23, 42, .12);
        }}

        .header {{
            background: linear-gradient(135deg, #1d6ff2 0%, #0f5db8 45%, #0d3f73 100%);
            color: #ffffff;
            padding: 28px;
        }}

        .header h1 {{
            margin: 0;
            font-size: 28px;
        }}

        .header p {{
            margin: 8px 0 0;
            opacity: .9;
        }}

        .content {{
            padding: 28px;
        }}

        h2 {{
            color: #0d3f73;
            margin-top: 0;
        }}

        table {{
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 28px;
            font-size: 14px;
        }}

        th {{
            background: #0f5db8;
            color: #ffffff;
            text-align: left;
            padding: 10px;
        }}

        td {{
            border: 1px solid #cbd5e1;
            padding: 10px;
        }}

        tr:nth-child(even) td {{
            background: #f8fafc;
        }}

        .footer {{
            border-top: 1px solid #e2e8f0;
            color: #64748b;
            font-size: 12px;
            padding: 18px 28px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>Dashboard Gerencial QuickStore</h1>
            <p>Reporte generado el {escape(generated_at)}</p>
        </div>

        <div class="content">
            <h2>Indicadores principales</h2>
            <table>
                <thead>
                    <tr>
                        <th>Indicador</th>
                        <th>Valor</th>
                    </tr>
                </thead>
                <tbody>
                    {indicator_html_rows}
                </tbody>
            </table>

            <h2>Ventas por día del mes actual</h2>
            <table>
                <thead>
                    <tr>
                        <th>Día</th>
                        <th>Fecha</th>
                        <th>Total vendido</th>
                    </tr>
                </thead>
                <tbody>
                    {sales_html_rows}
                </tbody>
            </table>
        </div>

        <div class="footer">
            Este reporte fue generado automáticamente por MiniMarket QuickStore.
        </div>
    </div>
</body>
</html>"""

        _register_dashboard_action(
            accion='DESCARGA DASHBOARD GERENCIAL HTML',
            detalle='El administrador descargó el Dashboard Gerencial en HTML.'
        )

        response = Response(html_content, mimetype='text/html')
        response.headers['Content-Disposition'] = f'attachment; filename={_get_report_filename("html")}'
        return response

    except Exception as e:
        _register_dashboard_action(
            accion='ERROR DESCARGA DASHBOARD GERENCIAL HTML',
            detalle=f'Error al descargar HTML del Dashboard Gerencial: {str(e)}'
        )

        flash(f'No se pudo generar el HTML: {str(e)}', 'danger')
        return redirect(url_for('managerial_dashboard.managerial_dashboard_page'))