from flask import Blueprint, render_template, request, send_file, session, flash, redirect, url_for
from models.abc_classification import ABCClassification
from utils.decorators import login_required, permission_required

from io import BytesIO
from datetime import date, datetime
from urllib.parse import urlencode

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer
)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


abc_classification_bp = Blueprint(
    'abc_classification',
    __name__,
    url_prefix='/reports/abc-classification'
)


def _clean_filter(value):
    """
    Limpiar filtros vacíos.
    """
    if value is None:
        return None

    value = str(value).strip()

    if value == '':
        return None

    return value


def _get_default_dates():
    """
    Fechas por defecto: desde el primer día del mes actual hasta hoy.
    """
    today = date.today()
    first_day = today.replace(day=1)

    return {
        'fecha_inicio': first_day.strftime('%Y-%m-%d'),
        'fecha_fin': today.strftime('%Y-%m-%d')
    }


def _get_filters_from_request():
    """
    Obtener filtros del request GET.
    El rango de fechas es obligatorio, pero se precarga con el mes actual.
    """
    default_dates = _get_default_dates()

    return {
        'fecha_inicio': _clean_filter(request.args.get('fecha_inicio')) or default_dates['fecha_inicio'],
        'fecha_fin': _clean_filter(request.args.get('fecha_fin')) or default_dates['fecha_fin'],
        'search': _clean_filter(request.args.get('search')),
        'classification': _clean_filter(request.args.get('classification'))
    }


def _build_export_query(filters):
    """
    Construir querystring para exportaciones PDF y Excel.
    """
    query_data = {}

    for key, value in filters.items():
        if value:
            query_data[key] = value

    return urlencode(query_data, doseq=True)


def _empty_summary():
    """
    Resumen vacío para cuando ocurre un error o no existen datos.
    """
    return {
        'total_productos_vendidos': 0,
        'total_productos_mostrados': 0,
        'total_unidades_vendidas': 0,
        'total_valor_vendido': 0,
        'productos_clase_a': 0,
        'productos_clase_b': 0,
        'productos_clase_c': 0,
        'valor_clase_a': 0,
        'valor_clase_b': 0,
        'valor_clase_c': 0
    }


def _format_money(value):
    """
    Formatear moneda para HTML/PDF.
    """
    try:
        return f'Bs {float(value):,.2f}'
    except Exception:
        return 'Bs 0.00'


def _format_number(value):
    """
    Formatear números enteros o decimales.
    """
    try:
        number = float(value)

        if number.is_integer():
            return f'{int(number):,}'

        return f'{number:,.2f}'
    except Exception:
        return '0'


def _format_percent(value):
    """
    Formatear porcentajes.
    """
    try:
        return f'{float(value):.2f}%'
    except Exception:
        return '0.00%'


def _build_pdf(rows, summary, filters):
    """
    Generar PDF del reporte ABC usando ReportLab.
    """
    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=1.1 * cm,
        leftMargin=1.1 * cm,
        topMargin=1.1 * cm,
        bottomMargin=1.1 * cm
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        name='TitleCenter',
        parent=styles['Title'],
        alignment=TA_CENTER,
        fontSize=16,
        leading=20,
        spaceAfter=8
    )

    subtitle_style = ParagraphStyle(
        name='SubtitleCenter',
        parent=styles['Normal'],
        alignment=TA_CENTER,
        fontSize=9,
        textColor=colors.HexColor('#555555'),
        spaceAfter=10
    )

    normal_style = ParagraphStyle(
        name='NormalABC',
        parent=styles['Normal'],
        fontSize=8,
        leading=10,
        alignment=TA_LEFT
    )

    small_style = ParagraphStyle(
        name='SmallABC',
        parent=styles['Normal'],
        fontSize=7,
        leading=9
    )

    elements = []

    elements.append(Paragraph('MiniMarket QuickStore', title_style))
    elements.append(Paragraph('Clasificación ABC de Productos', title_style))

    generated_at = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    elements.append(Paragraph(f'Generado el: {generated_at}', subtitle_style))

    filtros_texto = [
        f"Fecha inicial: {filters.get('fecha_inicio')}",
        f"Fecha final: {filters.get('fecha_fin')}"
    ]

    if filters.get('search'):
        filtros_texto.append(f"Búsqueda: {filters.get('search')}")

    if filters.get('classification'):
        filtros_texto.append(f"Clasificación: {filters.get('classification')}")

    elements.append(Paragraph('Filtros: ' + ' | '.join(filtros_texto), normal_style))
    elements.append(Spacer(1, 0.25 * cm))

    resumen_data = [
        [
            'Productos vendidos',
            'Unidades vendidas',
            'Valor vendido',
            'Clase A',
            'Clase B',
            'Clase C'
        ],
        [
            str(summary.get('total_productos_vendidos', 0)),
            _format_number(summary.get('total_unidades_vendidas', 0)),
            _format_money(summary.get('total_valor_vendido', 0)),
            str(summary.get('productos_clase_a', 0)),
            str(summary.get('productos_clase_b', 0)),
            str(summary.get('productos_clase_c', 0))
        ]
    ]

    resumen_table = Table(resumen_data, repeatRows=1)
    resumen_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#cbd5e1')),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#eff6ff')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
    ]))

    elements.append(resumen_table)
    elements.append(Spacer(1, 0.35 * cm))

    header = [
        Paragraph('<b>Código</b>', small_style),
        Paragraph('<b>Producto</b>', small_style),
        Paragraph('<b>Cantidad vendida</b>', small_style),
        Paragraph('<b>Valor vendido</b>', small_style),
        Paragraph('<b>% participación</b>', small_style),
        Paragraph('<b>% acumulado</b>', small_style),
        Paragraph('<b>Clase</b>', small_style),
    ]

    data = [header]

    for row in rows:
        data.append([
            Paragraph(str(row.get('codigo_producto') or '-'), small_style),
            Paragraph(str(row.get('producto') or '-'), small_style),
            Paragraph(_format_number(row.get('cantidad_vendida')), small_style),
            Paragraph(_format_money(row.get('valor_vendido')), small_style),
            Paragraph(_format_percent(row.get('porcentaje_participacion')), small_style),
            Paragraph(_format_percent(row.get('porcentaje_acumulado')), small_style),
            Paragraph(str(row.get('clasificacion') or '-'), small_style),
        ])

    if len(data) == 1:
        data.append([
            Paragraph('No existen datos para los filtros seleccionados.', small_style),
            '',
            '',
            '',
            '',
            '',
            ''
        ])

    table = Table(
        data,
        repeatRows=1,
        colWidths=[
            2.1 * cm,
            7.2 * cm,
            3.0 * cm,
            3.0 * cm,
            3.0 * cm,
            3.0 * cm,
            1.6 * cm
        ]
    )

    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#cbd5e1')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
        ('ALIGN', (6, 1), (6, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [
            colors.white,
            colors.HexColor('#f8fafc')
        ]),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
    ]))

    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    return buffer


def _build_excel(rows, summary, filters):
    """
    Generar archivo Excel del reporte ABC.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Clasificación ABC'

    header_fill = PatternFill('solid', fgColor='1E3A8A')
    subheader_fill = PatternFill('solid', fgColor='0F172A')
    white_font = Font(color='FFFFFF', bold=True)
    title_font = Font(size=16, bold=True, color='1E3A8A')
    subtitle_font = Font(size=11, color='475569')
    border_color = 'CBD5E1'
    thin_border = Border(
        left=Side(style='thin', color=border_color),
        right=Side(style='thin', color=border_color),
        top=Side(style='thin', color=border_color),
        bottom=Side(style='thin', color=border_color)
    )

    sheet.merge_cells('A1:G1')
    sheet['A1'] = 'MiniMarket QuickStore - Clasificación ABC de Productos'
    sheet['A1'].font = title_font
    sheet['A1'].alignment = Alignment(horizontal='center')

    sheet.merge_cells('A2:G2')
    sheet['A2'] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    sheet['A2'].font = subtitle_font
    sheet['A2'].alignment = Alignment(horizontal='center')

    sheet['A4'] = 'Fecha inicial'
    sheet['B4'] = filters.get('fecha_inicio')
    sheet['C4'] = 'Fecha final'
    sheet['D4'] = filters.get('fecha_fin')
    sheet['E4'] = 'Búsqueda'
    sheet['F4'] = filters.get('search') or '-'
    sheet['G4'] = f"Clase: {filters.get('classification') or 'Todas'}"

    for cell in sheet[4]:
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    summary_headers = [
        'Productos vendidos',
        'Unidades vendidas',
        'Valor vendido',
        'Clase A',
        'Clase B',
        'Clase C'
    ]

    summary_values = [
        summary.get('total_productos_vendidos', 0),
        float(summary.get('total_unidades_vendidas', 0)),
        float(summary.get('total_valor_vendido', 0)),
        summary.get('productos_clase_a', 0),
        summary.get('productos_clase_b', 0),
        summary.get('productos_clase_c', 0),
    ]

    start_summary_row = 6

    for col, value in enumerate(summary_headers, start=1):
        cell = sheet.cell(row=start_summary_row, column=col, value=value)
        cell.fill = header_fill
        cell.font = white_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    for col, value in enumerate(summary_values, start=1):
        cell = sheet.cell(row=start_summary_row + 1, column=col, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    table_start = 10

    headers = [
        'Código',
        'Producto',
        'Cantidad vendida',
        'Valor vendido',
        '% participación',
        '% acumulado',
        'Clasificación ABC'
    ]

    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=table_start, column=col, value=header)
        cell.fill = subheader_fill
        cell.font = white_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    current_row = table_start + 1

    if rows:
        for row in rows:
            values = [
                row.get('codigo_producto'),
                row.get('producto'),
                float(row.get('cantidad_vendida') or 0),
                float(row.get('valor_vendido') or 0),
                float(row.get('porcentaje_participacion') or 0) / 100,
                float(row.get('porcentaje_acumulado') or 0) / 100,
                row.get('clasificacion')
            ]

            for col, value in enumerate(values, start=1):
                cell = sheet.cell(row=current_row, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal='left' if col == 2 else 'center'
                )

                if col in [4]:
                    cell.number_format = '#,##0.00'

                if col in [5, 6]:
                    cell.number_format = '0.00%'

            current_row += 1
    else:
        sheet.cell(
            row=current_row,
            column=1,
            value='No existen datos para los filtros seleccionados.'
        )
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)

    widths = {
        1: 14,
        2: 38,
        3: 18,
        4: 18,
        5: 18,
        6: 18,
        7: 18
    }

    for col, width in widths.items():
        sheet.column_dimensions[get_column_letter(col)].width = width

    sheet.freeze_panes = 'A11'

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    return buffer


@abc_classification_bp.route('/')
@login_required
@permission_required('report_profit_read')
def abc_report():
    """
    Pantalla principal del reporte de Clasificación ABC.
    """
    filters = _get_filters_from_request()
    rows = []
    summary = _empty_summary()

    try:
        result = ABCClassification.get_report(
            fecha_inicio=filters['fecha_inicio'],
            fecha_fin=filters['fecha_fin'],
            search=filters['search'],
            classification=filters['classification']
        )

        rows = result['rows']
        summary = result['summary']

        ABCClassification.register_action(
            codigo_usuario=session.get('user_id'),
            accion='CONSULTA CLASIFICACIÓN ABC',
            detalle=(
                f"Consulta ABC desde {filters['fecha_inicio']} hasta {filters['fecha_fin']} "
                f"con {summary['total_productos_mostrados']} productos mostrados."
            )
        )

    except ValueError as e:
        flash(str(e), 'danger')

    except Exception as e:
        flash(f'Error al generar la clasificación ABC: {str(e)}', 'danger')

    export_query = _build_export_query(filters)

    return render_template(
        'reports/abc_classification.html',
        rows=rows,
        summary=summary,
        filters=filters,
        export_query=export_query,
        format_money=_format_money,
        format_number=_format_number,
        format_percent=_format_percent
    )


@abc_classification_bp.route('/pdf')
@login_required
@permission_required('report_profit_read')
def abc_report_pdf():
    """
    Descargar Clasificación ABC en PDF.
    """
    filters = _get_filters_from_request()

    try:
        result = ABCClassification.get_report(
            fecha_inicio=filters['fecha_inicio'],
            fecha_fin=filters['fecha_fin'],
            search=filters['search'],
            classification=filters['classification']
        )

        rows = result['rows']
        summary = result['summary']

        pdf_buffer = _build_pdf(
            rows=rows,
            summary=summary,
            filters=filters
        )

        ABCClassification.register_action(
            codigo_usuario=session.get('user_id'),
            accion='DESCARGA CLASIFICACIÓN ABC PDF',
            detalle=(
                f"Descarga PDF ABC desde {filters['fecha_inicio']} hasta {filters['fecha_fin']} "
                f"con {len(rows)} productos."
            )
        )

        filename = f'clasificacion_abc_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'

        return send_file(
            pdf_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )

    except ValueError as e:
        flash(str(e), 'danger')

    except Exception as e:
        flash(f'Error al descargar PDF de clasificación ABC: {str(e)}', 'danger')

    return redirect(url_for('abc_classification.abc_report'))


@abc_classification_bp.route('/excel')
@login_required
@permission_required('report_profit_read')
def abc_report_excel():
    """
    Descargar Clasificación ABC en Excel.
    """
    filters = _get_filters_from_request()

    try:
        result = ABCClassification.get_report(
            fecha_inicio=filters['fecha_inicio'],
            fecha_fin=filters['fecha_fin'],
            search=filters['search'],
            classification=filters['classification']
        )

        rows = result['rows']
        summary = result['summary']

        excel_buffer = _build_excel(
            rows=rows,
            summary=summary,
            filters=filters
        )

        ABCClassification.register_action(
            codigo_usuario=session.get('user_id'),
            accion='DESCARGA CLASIFICACIÓN ABC EXCEL',
            detalle=(
                f"Descarga Excel ABC desde {filters['fecha_inicio']} hasta {filters['fecha_fin']} "
                f"con {len(rows)} productos."
            )
        )

        filename = f'clasificacion_abc_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except ValueError as e:
        flash(str(e), 'danger')

    except Exception as e:
        flash(f'Error al descargar Excel de clasificación ABC: {str(e)}', 'danger')

    return redirect(url_for('abc_classification.abc_report'))